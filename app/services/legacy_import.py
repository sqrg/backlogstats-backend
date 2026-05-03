"""Parse the user's legacy backlog xlsx into structured rows.

Used by the `/api/v1/import/legacy/parse` endpoint. Pure function — no DB,
no IGDB, just sheet → list of dicts shaped for `ParsedLegacyRow`.

Source layout (sheets named by year, leading blank column A):
- 2020, 2021: Juego, Plataforma, Horas, Fecha de fin
- 2022:      Título, Es DLC?, Jugado en portátil?, Plataforma, Horas HLTB,
             Horas, Fecha de fin
- 2023+:     adds Horas este año (ignored)

`Horas HLTB` and `Horas este año` are intentionally ignored. Stats / formula
columns to the right of the data block are also ignored — we only read the
columns named in `_row_columns`.
"""

from __future__ import annotations

import io
from datetime import date, datetime

import openpyxl

# Maps every distinct platform string seen in the spreadsheet to the canonical
# `Platform.name` we want in the DB.
#
# Notes:
#  - "One X" alone and the ambiguous "One X / Series X" both default to
#    Xbox One. Pure "Series X" stays Xbox Series X.
#  - "VR" maps to PC; the user can override per-row in the UI.
#  - "Wii U" stays distinct from "Wii".
PLATFORM_SYNONYMS: dict[str, str] = {
    "PC": "PC",
    "Switch": "Nintendo Switch",
    "Nintendo Switch": "Nintendo Switch",
    "PS Vita": "PlayStation Vita",
    "Vita": "PlayStation Vita",
    "PS1": "PlayStation",
    "PS3": "PlayStation 3",
    "PS4": "PlayStation 4",
    "PS5": "PlayStation 5",
    "PSP": "PlayStation Portable",
    "Xbox 360": "Xbox 360",
    "X360": "Xbox 360",
    "One X": "Xbox One",
    "One X / Series X": "Xbox One",
    "Series X": "Xbox Series X",
    "GBA": "Game Boy Advance",
    "Gameboy": "Game Boy",
    "GC": "Nintendo GameCube",
    "Gamecube": "Nintendo GameCube",
    "SNES": "Super Nintendo Entertainment System",
    "NES": "Nintendo Entertainment System",
    "N64": "Nintendo 64",
    "DS": "Nintendo DS",
    "3DS": "Nintendo 3DS",
    "Wii": "Wii",
    "Wii U": "Wii U",
    "Genesis": "Sega Genesis",
    "Dreamcast": "Dreamcast",
    "Mobile": "Mobile",
    "VR": "PC",
}


def _normalize_platform(raw: str | None) -> tuple[str, str | None]:
    if not raw:
        return ("Unknown", "platform missing in source row")
    cleaned = raw.strip()
    canonical = PLATFORM_SYNONYMS.get(cleaned)
    if canonical is None:
        return (cleaned, f"unmapped platform '{raw}', kept verbatim")
    note: str | None = None
    if cleaned == "One X / Series X":
        note = "ambiguous in source: One X / Series X (defaulted to Xbox One)"
    elif cleaned == "VR":
        note = "source platform was 'VR', mapped to PC"
    return (canonical, note)


def _coerce_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _coerce_hours(value: object) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _coerce_bool_es(value: object) -> bool:
    """Source uses 'Si'/'No' (with stray 'Yes'/'Ambas'/'Both' since 2023)."""
    if value is None:
        return False
    s = str(value).strip().lower()
    return s in {"si", "sí", "yes", "y", "true", "ambas", "both"}


def _row_columns(year: int) -> dict[str, int]:
    """Column indices (0-based) for the data block on each sheet."""
    if year in (2020, 2021):
        return {"title": 1, "platform": 2, "hours": 3, "completed_at": 4}
    if year == 2022:
        return {
            "title": 1,
            "is_dlc": 2,
            "is_handheld": 3,
            "platform": 4,
            "hours": 6,
            "completed_at": 7,
        }
    return {
        "title": 1,
        "is_dlc": 2,
        "is_handheld": 3,
        "platform": 4,
        "hours": 6,
        "completed_at": 8,
    }


def parse_legacy_xlsx(file_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    out: list[dict] = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.isdigit():
            continue
        year = int(sheet_name)
        ws = wb[sheet_name]
        cols = _row_columns(year)

        for excel_row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            title = row[cols["title"]] if cols["title"] < len(row) else None
            if not title:
                continue
            platform_raw = (
                row[cols["platform"]] if cols["platform"] < len(row) else None
            )
            hours = _coerce_hours(
                row[cols["hours"]] if cols["hours"] < len(row) else None
            )
            completed_at = _coerce_date(
                row[cols["completed_at"]]
                if cols["completed_at"] < len(row)
                else None
            )
            is_dlc = (
                _coerce_bool_es(row[cols["is_dlc"]])
                if "is_dlc" in cols and cols["is_dlc"] < len(row)
                else False
            )
            is_handheld = (
                _coerce_bool_es(row[cols["is_handheld"]])
                if "is_handheld" in cols and cols["is_handheld"] < len(row)
                else False
            )

            canonical_platform, platform_note = _normalize_platform(
                str(platform_raw) if platform_raw else None
            )

            out.append(
                {
                    "row_id": f"{year}:{excel_row_idx}",
                    "title_raw": str(title).strip(),
                    "platform_raw": (
                        str(platform_raw).strip() if platform_raw else None
                    ),
                    "platform_normalized": canonical_platform,
                    "platform_note": platform_note,
                    "hours": hours,
                    "completed_at": completed_at,
                    "is_dlc": is_dlc,
                    "is_handheld": is_handheld,
                    "year_sheet": year,
                }
            )

    return out
